# Packages and Modules
from linkedlist import LinkedList
import openpyxl as xl

# Define Bubblesort implementation using Linked List
def linkbubblesort(linkList):
    for i in range(len(linkList)-1, 0, -1):
        thead = linkList._head
        for j in range(i):
            if thead._element > thead._next._element:
                thead._element, thead._next._element = thead._next._element, thead._element
            thead = thead._next
# Passes of the bubble sort method:
        linkList.display()

# loading workbook into the code or program
wb = xl.load_workbook("pythonxlsx.xlsx")
sheet = wb["Sheet1"]

LL = LinkedList()

# for loop for reading the cells in the xlsx file
for row in range(1, sheet.max_row + 1):
    cell = sheet.cell(row, 1)
    LL.add_last(cell.value)

LL.display()
# Sorting the Linked List
linkbubblesort(LL)
# This loop should take in the data from the linked list and import it into the rows of column 2
row = 1
thead=LL._head
while thead:
    cell_sorted = sheet.cell(row, 3)
    print(thead._element)
    cell_sorted.value = thead._element
    row += 1
    thead = thead._next
wb.save("d:\\ExcelSheet.xlsx")
print("Done")
# i wrote this whole thing
