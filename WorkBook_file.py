import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    print("Start process_workbook")
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row,3)
        print(cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values=Reference (sheet,
               min_row=2,
               max_row=sheet.max_row,
               min_col=4,
               max_col=4)

    chart=BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, "g5")

    wb.save("har1.xlsx")

print("Start")
process_workbook("D:\\Harshith Extras\\HArshith Codes\\HelloWorld\\transactions.xlsx")
print("End")